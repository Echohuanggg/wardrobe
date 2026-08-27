# -*- coding: utf-8 -*-
"""
衣物清单同步脚本：以 Excel 为权威源，自动更新 wardrobe.json 与 thumbnails 缩略图。
用法：python sync_excel_to_json.py
查找顺序：1) 脚本同目录下的 衣物识别清单.xlsx  2) 上级目录下的 衣物识别清单.xlsx
"""
import json
import os
import re
import shutil
import sys
import zipfile
from datetime import date

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
EXCEL_NAMES = ["衣物识别清单.xlsx"]
OUT_JSON = os.path.join(BASE_DIR, "wardrobe.json")
THUMB_DIR = os.path.join(BASE_DIR, "thumbnails")
BAK_JSON = os.path.join(BASE_DIR, "wardrobe.json.bak")


def find_excel():
    """依次在脚本目录、上级目录查找 Excel"""
    for base in (BASE_DIR, os.path.dirname(BASE_DIR)):
        for name in EXCEL_NAMES:
            p = os.path.join(base, name)
            if os.path.exists(p):
                return p
    return None


def read_excel_rows(excel_path):
    """用 zipfile 方式解析 xlsx：读取共享字符串 + 单元格文本，避免依赖 openpyxl"""
    try:
        from openpyxl import load_workbook
        return read_with_openpyxl(excel_path)
    except Exception:
        return read_with_zip(excel_path)


def read_with_openpyxl(excel_path):
    from openpyxl import load_workbook
    wb = load_workbook(excel_path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        return []
    header = [str(h).strip() if h is not None else "" for h in rows[0]]
    col = {name: i for i, name in enumerate(header)}
    items = []
    for r in rows[1:]:
        if r[0] is None:
            continue
        try:
            rid = int(r[0])
        except (ValueError, TypeError):
            continue
        items.append({
            "id": rid,
            "category": str(r[col["类别"]] or "").strip() if "类别" in col else "",
            "color": str(r[col["颜色"]] or "").strip() if "颜色" in col else "",
            "style": str(r[col["款式"]] or "").strip() if "款式" in col else "",
            "brand": str(r[col["品牌"]] or "").strip() if "品牌" in col else "",
            "material": str(r[col["材质"]] or "").strip() if "材质" in col else "",
            "purchase_date": str(r[col["购买日期"]] or "").strip() if "购买日期" in col else "",
            "seasons_raw": str(r[col["季节"]] or "").strip() if "季节" in col else "",
            "note": str(r[col["备注"]] or "").strip() if "备注" in col else "",
        })
    return items


def read_with_zip(excel_path):
    """降级方案：直接解析 sheet XML + sharedStrings，不依赖第三方库"""
    items = []
    try:
        with zipfile.ZipFile(excel_path) as z:
            shared = []
            if "xl/sharedStrings.xml" in z.namelist():
                sxml = z.read("xl/sharedStrings.xml").decode("utf-8", errors="ignore")
                # 逐 si 提取全部 <t> 文本拼接（兼容富文本与 <phoneticPr/> 干扰）
                for si in re.findall(r"<si>(.*?)</si>", sxml, re.S):
                    texts = re.findall(r"<t[^>]*>([^<]*)</t>", si)
                    shared.append("".join(texts) if texts else "")
            sheet_name = [n for n in z.namelist() if re.match(r"xl/worksheets/sheet\d+\.xml", n)][0]
            xml = z.read(sheet_name).decode("utf-8", errors="ignore")
            rows_data = []  # [(excel_row, {col: val})]
            # 按行解析
            for row_m in re.finditer(r"<row[^>]*r=\"(\d+)\"[^>]*>(.*?)</row>", xml, re.S):
                excel_row = int(row_m.group(1))
                cells = {}
                # 整体捕获属性串，再从中提取 t 类型（避免非贪婪+可选组导致的 t 丢失）
                for c in re.finditer(r'<c r="([A-Z]+)\d+"([^>]*)>(?:<v>([^<]*)</v>)?(?:<is><t[^>]*>([^<]*)</t></is>)?</c>', row_m.group(2)):
                    col_letter, attrs, v, inline = c.group(1), c.group(2), c.group(3), c.group(4)
                    val = ""
                    if inline is not None:
                        val = inline
                    else:
                        tm = re.search(r'\bt="([^"]+)"', attrs)
                        if tm and tm.group(1) == "s" and v is not None:
                            val = shared[int(v)]
                        elif v is not None:
                            val = v
                    cells[col_letter] = val.strip()
                rows_data.append((excel_row, cells))
            if not rows_data:
                return items
            # 表头行：列名 -> 列字母（增列后仍能正确对应）
            header_cells = dict(rows_data[0][1])
            col = {}
            for letter, name in header_cells.items():
                col[str(name).strip()] = letter
            for excel_row, cells in rows_data[1:]:
                try:
                    rid = int(cells.get(col.get("编号", "A"), ""))
                except (ValueError, TypeError):
                    continue
                items.append({
                    "id": rid,
                    "category": cells.get(col.get("类别", "B"), ""),
                    "color": cells.get(col.get("颜色", "C"), ""),
                    "style": cells.get(col.get("款式", "D"), ""),
                    "brand": cells.get(col.get("品牌", "E"), ""),
                    "material": cells.get(col.get("材质", "F"), ""),
                    "purchase_date": cells.get(col.get("购买日期", "G"), ""),
                    "seasons_raw": cells.get(col.get("季节", "H"), ""),
                    "note": cells.get(col.get("备注", "I"), ""),
                })
    except Exception as e:
        print(f"[错误] 解析 Excel 失败: {e}")
        sys.exit(1)
    return items


def extract_images(excel_path):
    """从 Excel 内嵌图导出缩略图：解析 drawing1.xml 锚点 -> 编号 -> xl/media/*.png"""
    exported = {}
    try:
        with zipfile.ZipFile(excel_path) as z:
            names = z.namelist()
            drawing = [n for n in names if re.match(r"xl/drawings/drawing\d+\.xml", n)]
            if not drawing:
                return exported
            xml = z.read(drawing[0]).decode("utf-8", errors="ignore")
            # rels 映射 rId -> media 文件（属性顺序不固定，逐块提取 Id/Target）
            rel_path = drawing[0].replace("drawings/", "drawings/_rels/") + ".rels"
            rels = {}
            if rel_path in names:
                rxml = z.read(rel_path).decode("utf-8", errors="ignore")
                for rm in re.finditer(r"<Relationship[^>]*/>", rxml):
                    rid_m = re.search(r'Id="([^"]+)"', rm.group(0))
                    tgt_m = re.search(r'Target="([^"]+)"', rm.group(0))
                    if rid_m and tgt_m:
                        rels[rid_m.group(1)] = tgt_m.group(1)
            # oneCellAnchor / twoCellAnchor 锚点（兼容 xdr: 前缀与默认命名空间、带属性标签）
            anchors = re.findall(r"<(?:xdr:)?(?:oneCellAnchor|twoCellAnchor)[^>]*>(.*?)</(?:xdr:)?(?:oneCellAnchor|twoCellAnchor)>", xml, re.S)
            for body in anchors:
                row_m = re.search(r"<(?:xdr:)?from>.*?<(?:xdr:)?row>(\d+)</(?:xdr:)?row>", body, re.S)
                embed_m = re.search(r'r:embed="([^"]+)"', body)
                if not row_m or not embed_m:
                    continue
                excel_row = int(row_m.group(1)) + 1  # row 是 0-based，+1 = Excel 行号
                rid = embed_m.group(1)
                target = rels.get(rid, "")
                media_file = target.lstrip("/").replace("../", "xl/")
                if media_file in names:
                    exported[excel_row] = media_file
    except Exception as e:
        print(f"[警告] 提取内嵌图失败: {e}")
    return exported


def parse_seasons(raw):
    """把季节文本转数组：'春夏秋'->['春','夏','秋']，'四季'->全季"""
    if not raw or raw == "无":
        return []
    if "四季" in raw or raw == "全年":
        return ["春", "夏", "秋", "冬"]
    result = []
    for s in "春夏秋冬":
        if s in raw:
            result.append(s)
    return result


def parse_purchase_date(raw):
    """购买日期：数字 202608 -> '202608'，空 -> ''"""
    if not raw:
        return ""
    if raw in ("无", "未知", "N/A"):
        return ""
    return str(raw).strip()


def main():
    excel_path = find_excel()
    if not excel_path:
        print("[错误] 未找到 衣物识别清单.xlsx（脚本目录或上级目录）")
        sys.exit(1)
    print(f"[1/4] 读取 Excel: {excel_path}")

    rows = read_excel_rows(excel_path)
    if not rows:
        print("[错误] Excel 无数据行")
        sys.exit(1)
    print(f"[2/4] Excel 共 {len(rows)} 件衣物（编号 {min(r['id'] for r in rows)}~{max(r['id'] for r in rows)}）")

    # 读取旧 JSON（保留 name/occasion/created_at 等字段）
    old_map = {}
    if os.path.exists(OUT_JSON):
        try:
            with open(OUT_JSON, encoding="utf-8") as f:
                old_map = {int(item["id"]): item for item in json.load(f)}
        except Exception:
            old_map = {}
    if old_map:
        shutil.copy2(OUT_JSON, BAK_JSON)
        print(f"    旧 JSON 已备份: {BAK_JSON}")

    # 提取内嵌图
    image_map = extract_images(excel_path)  # Excel行号 -> media文件
    print(f"[3/4] Excel 内嵌缩略图 {len(image_map)} 张")

    os.makedirs(THUMB_DIR, exist_ok=True)
    new_items = []
    exported_count = 0
    for r in sorted(rows, key=lambda x: x["id"]):
        rid = r["id"]
        old = old_map.get(rid, {})
        name = old.get("name") or f"{r['color']}{r['style']}".strip() or f"衣物{rid:03d}"
        # 品牌归一化
        brand = r["brand"]
        if brand == "无":
            brand = ""
        seasons = parse_seasons(r["seasons_raw"])
        purchase_date = parse_purchase_date(r["purchase_date"])
        # 缩略图：优先从 Excel 内嵌图导出
        img_file = f"thumbnails/{rid:03d}.png"
        media = image_map.get(rid)
        if media:
            try:
                with zipfile.ZipFile(excel_path) as z:
                    data = z.read(media)
                with open(os.path.join(BASE_DIR, img_file), "wb") as f:
                    f.write(data)
                exported_count += 1
            except Exception as e:
                print(f"    [警告] 导出 #{rid} 图片失败: {e}")
        elif not os.path.exists(os.path.join(BASE_DIR, img_file)):
            print(f"    [提示] #{rid} 无内嵌图且无现有缩略图，生成白底占位图")
            _make_placeholder(os.path.join(BASE_DIR, img_file))

        item = {
            "id": rid,
            "name": name,
            "category": r["category"] or old.get("category", ""),
            "color": r["color"] or old.get("color", ""),
            "style": r["style"] or old.get("style", ""),
            "brand": brand or old.get("brand", ""),
            "material": r["material"] or old.get("material", ""),
            "seasons": seasons if seasons else old.get("seasons", []),
            "occasion": old.get("occasion", ""),
            "purchase_date": purchase_date if purchase_date else old.get("purchase_date", ""),
            "note": r["note"] or old.get("note", ""),
            "created_at": old.get("created_at", date.today().isoformat()),
            "image": img_file,
        }
        new_items.append(item)

    # 写 JSON
    with open(OUT_JSON, "w", encoding="utf-8") as f:
        json.dump(new_items, f, ensure_ascii=False, indent=2)

    # 验证
    with open(OUT_JSON, encoding="utf-8") as f:
        check = json.load(f)
    missing_img = [i["id"] for i in check if not os.path.exists(os.path.join(BASE_DIR, i["image"]))]
    print(f"[4/4] 同步完成！wardrobe.json 共 {len(check)} 条，导出/刷新缩略图 {exported_count} 张")
    if missing_img:
        print(f"    [警告] 以下编号缺缩略图文件: {missing_img}")
    else:
        print("    所有缩略图文件均可访问")
    print("数据已就绪，可以启动页面。")


def _make_placeholder(path):
    """生成 160x160 白底 PNG 占位图（纯 Python，无 PIL 依赖）"""
    try:
        from PIL import Image
        img = Image.new("RGB", (160, 160), (255, 255, 255))
        img.save(path)
    except Exception:
        # 无 PIL 时写最小 PNG
        import struct, zlib
        w = h = 160
        raw = b"".join(b"\x00" + b"\xff\xff\xff" * w for _ in range(h))
        def chunk(t, d):
            c = t + d
            return struct.pack(">I", len(d)) + c + struct.pack(">I", zlib.crc32(c))
        png = b"\x89PNG\r\n\x1a\n"
        png += chunk(b"IHDR", struct.pack(">IIBBBBB", w, h, 8, 2, 0, 0, 0))
        png += chunk(b"IDAT", zlib.compress(raw))
        png += chunk(b"IEND", b"")
        with open(path, "wb") as f:
            f.write(png)


if __name__ == "__main__":
    main()
